#実運用のルールのバックテスト(excel.csvデータを入力、出力)
#ルール　①Hyperによる取引(日中と夜の立会ごとに決済）　②値幅30円で更新　③前の山(谷)を超えて+30で新規、前の谷(山)を超えて-20で決済　④含み損100で損切り　⑤手数料は売買1回で100円と仮定 ⑥1年分でテスト(sq日ごとに計算はしない)
import os
import pandas as pd
import openpyxl as px
from openpyxl.chart import LineChart,Reference,series
import datetime
from collections import deque

#df_2 = [日付、時間、始値、高値、安値、終値](日付、時間はdatetime型(時間をtime型にしたかったけどよく分からなかった))
#夜相場の閉め時のデータは削除、次の関数でindexが倍数になるごとにリストを作るためにキリ良くしたかった
def get1minsheet():
    time_1 = datetime.datetime(1900,1,1,5,30)

    df_1 = pd.read_csv(os.getcwd()+"\\N225minif_2019.csv",skiprows=1,names=("Date","Time","Open","High","Low","Close"),encoding='cp932',usecols=["Date","Time","Open","High","Low","Close"])

    df_1["Date"] = pd.to_datetime(df_1["Date"],format='%Y/%m/%d')
    df_1["Time"] = pd.to_datetime(df_1["Time"],format='%H:%M')

    df_2 = df_1[df_1["Time"] != time_1]
    
    return(df_2.reset_index(drop=True))

#ohlc_5 = [日付、時間、始値、[高値、安値](1分足)*tf、終値]⇒([0]:日付,[1]:時間,[2]:始値,[3]~[7]:高値、安値,[8]:終値)
def make_ohlc(df,tf):
    ohlc = []

    #append:１つの要素を追加可能、複数加える場合は2次元内包になる　extend:1度に複数要素を追加可能
    for num in range(len(df.index)-1):
        if num % tf == 0:
            ohlc_5 = []
            ohlc_5.extend([df.iat[num,0],df.iat[num,1],df.iat[num,2]])

        ohlc_5.append([df.iat[num,3],df.iat[num,4]])
        
        if num % tf == tf-1:
            ohlc_5.append(df.iat[num,5])
            ohlc.append(ohlc_5)

    return(ohlc)

#日中立会内での暫定高(安)値とその時の日時を記録する関数
def get225Data():

    sheet = make_ohlc(get1minsheet(),5)

    #[]:listを初期化する
    data = []
    prev = 0
    yori_time_am = datetime.datetime(1900,1,1,8,45) 
    hike_time_am = datetime.datetime(1900,1,1,15,15)
    yori_time_pm = datetime.datetime(1900,1,1,16,30)
    hike_time_pm = datetime.datetime(1900,1,1,5,30)

    #cell(1列目)に数値がなくなるまで繰り返す関数
    for i in sheet:
            #立会の時間帯か確認
            if i[1] < hike_time_am or yori_time_pm <= i[1]:
                #場の開始時に始値と暫定高値の幅が30円(窓開け)の場合
                if yori_time_am == i[1] and abs(i[2] - prev) >= 30 or yori_time_pm == i[1] and abs(i[2] - prev) >= 30:
                    #真：始値を暫定高(安)値として、その日時と値段をlistに追加、終値を決める(prev:暫定高(安)値、current:終値)　偽：終値を決める
                    prev = i[2]
                    data.append(i)
                
                current = i[8]

                #暫定高(安)値と終値の幅が30円以上の場合
                if abs(current - prev) >= 30:
                    #終値を暫定高(安)値として、日時と値段をlistに追加
                    prev = current
                    data.append(i)

    print("損益表名を入力してください ->")
    input_data_kagi = input()

    entry_point_kagi = []
    kagiSignal(data,entry_point_kagi)

    make_performance(entry_point_kagi,input_data_kagi+".xlsx")

def kagiSignal(list,entry_point):
    list_count = 0
    retu_count = 1
    #prev_listには暫定高(安)値とトレンド方向を記録,[0]=価格,[1]=トレンド方向
    prev_list = []
    #dequeは先頭、末尾どちらからも取り出せる(キュー、スタック(キューの取り出しはpopleft())
    #yama_tini_queueは直近の山谷形成価格を記録(高値、安値の2つ以上は存在しない)
    yama_tini_queue = deque([])

    #for 変数　in --: --の要素を変数として順番にする
    for w in list:
        #dataの要素1-2で上昇か下降か判断、暫定高(安)値をprev_list1に記録
        if list_count == 0 and list[0][8] == list[0][8]:
            if list_count == 0 and list[0][2] > list[1][8]:
                prev_list.extend([list[1][8],-1])
            else:
                prev_list.extend([list[1][8],1])
        elif list_count == 0 and list[0][8] > list[1][8]:
            prev_list.extend([list[1][8],-1])
        elif list_count == 0 and list[0][8] < list[1][8]:
            prev_list.extend([list[1][8],1])

        #記録されたトレンド方向に記録価格が進んでいたら暫定高(安)値更新
        if list_count >= 2:
            if prev_list[1] == 1 and prev_list[0] < w[8]:
                prev_list[0] = w[8]
                prev_list[1] = 1
            elif prev_list[1] == -1 and prev_list[0] > w[8]:
                prev_list[0] = w[8]
                prev_list[1] = -1
            #トレンドと逆方向に記録価格が進んだ場合
            elif prev_list[1] == 1 and prev_list[0] > w[2]:
                #記録されている直近の山谷形成価格が2つある場合、古い記録を削除し、新しい記録に更新(削除されるのは更新される価格と同じ向き)
                #ない場合は追加
                if len(yama_tini_queue) == 2:
                    yama_tini_queue.popleft()
                    yama_tini_queue.append(prev_list)
                else:
                    yama_tini_queue.append(prev_list)

                prev_list = [w[2],-1]

                retu_count += 1

            elif prev_list[1] == -1 and prev_list[0] < w[2]:
                if len(yama_tini_queue) == 2:
                    yama_tini_queue.popleft()
                    yama_tini_queue.append(prev_list)
                else:
                    yama_tini_queue.append(prev_list)
                
                prev_list = [w[2],1]

                retu_count += 1

            #yama_tini_queueが2つ記録しているとき、前の記録を取り出し現在価格、記録価格、トレンド方向、価格差、リスト数を表示
            if len(yama_tini_queue) == 2:
                temp = yama_tini_queue.popleft()
                print(w[8],temp[0],temp[1],w[8]-temp[0],list_count)

                #山谷超えのデータをtemp3(価格、日付、シグナル)としてnuki_pointに記録
                #超えなかった場合はyama_tini_queueに戻す
                if temp[1] == 1 and w[8] - temp[0] > 0:
                    temp3 = [w[0],w[1],w[8],"L",retu_count]
                    entry_point.append(temp3)
                elif temp[1] == -1 and temp[0] - w[8] > 0:
                    temp3 = [w[0],w[1],w[8],"S",retu_count]
                    entry_point.append(temp3)
                else:
                    yama_tini_queue.appendleft(temp)
        list_count += 1

def make_performance(entry_point,save_name):
    wb2 =px.Workbook()
    sheet2 = wb2.active
    sheet2.title =  "kagiashi"
    sheet2["A1"] = "エントリー日付"
    sheet2["B1"] = "エントリー"
    sheet2["C1"] = "決済日付"
    sheet2["D1"] = "決済"
    sheet2["E1"] = "シグナル"
    sheet2["F1"] = "損益"
    sheet2["G1"] = "損益合計"

    sum_sonneki = 0
    column_count = 2
    #sonneki内の-1は手数料代100円を想定
    for i in range(1,len(entry_point)):
        if entry_point[i-1][3] == 1:
            sonneki = (entry_point[i-1][2] - entry_point[i][2]-1)*100
        else:
            sonneki = (entry_point[i][2] - entry_point[i-1][2]-1)*100

        sum_sonneki += sonneki

        sheet2.cell(row=column_count, column=1, value=entry_point[i-1][0])
        sheet2.cell(row=column_count, column=2, value=entry_point[i-1][1])
        sheet2.cell(row=column_count, column=3, value=entry_point[i][0])
        sheet2.cell(row=column_count, column=4, value=entry_point[i][2])
        sheet2.cell(row=column_count, column=5, value=entry_point[i-1][3])
        sheet2.cell(row=column_count, column=6, value=sonneki)
        sheet2.cell(row=column_count, column=7, value=sum_sonneki)

        column_count += 1

    #折れ線グラフの作成
    values = Reference(sheet2,min_col=7,min_row=2,max_col=7,max_row=len(entry_point)+1)
    chart = LineChart()
    chart.add_data(values)
    sheet2.add_chart(chart,"I2")

    wb2.save(save_name)

get225Data()