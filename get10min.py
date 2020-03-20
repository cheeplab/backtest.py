#10分足で作成したカギ足によるバックテスト(ネットに挙がってたものを参考に作成)
　#ルール？　①日中立会内で売買　②値幅50円で更新　③前の山(谷)を超えたら新規、前の谷(山)を超えたら決済
import os
import openpyxl as px
import datetime
from collections import deque

def get10minSheet():
    print(os.getcwd())
    
    print("データをインプットしてください ->")
    input_data = input()

    print(os.getcwd()+"\\"+input_data)

    wb = px.load_workbook(os.getcwd()+"\\"+input_data)

    return wb["10min"]

#日中立会内での暫定高(安)値とその時の日時を記録する関数
def get225Data():
    
    sheet = get10minSheet()

    #[]:listを初期化する
    list1 = []
    list2 = []
    list3 = []
    list4 = []
    i = 2
    count = -1
    yori_frag = False
    prev1 = 0
    prev2 = 0
    prev3 = 0
    prev4 = 0

    yori_time = datetime.time(8,50,0) 
    hike_time = datetime.time(15,0,0)
    change_sq = datetime.date(2019,1,4)
    change_3sq = datetime.date(2019,3,1)
    change_6sq = datetime.date(2019,6,1)
    change_9sq = datetime.date(2019,9,1)
    chenge_12sq = datetime.date(2019,12,1)

    #cell(1列目)に数値がなくなるまで繰り返す関数
    while True:
        if not sheet.cell(row=i,column=1).value:
            break
        
        #1~3月の場であることを確認
        if 1 <= sheet.cell(row=i,column=1).value.month and sheet.cell(row=i,column=1).value.month <= 3:
            if yori_time == sheet.cell(row=i,column=2).value:
                yori_frag = True

            #日中立会の時間帯か確認
            if yori_frag == True and sheet.cell(row=i,column=2).value <= hike_time:
                #場の開始時に始値と暫定高値の幅が50円(窓開け)の場合
                if yori_time == sheet.cell(row=i,column=2).value and abs(sheet.cell(row=i,column=3).value - prev1) >= 50:
                    #真：始値を暫定高(安)値として、その日時と値段をlistに追加、終値を決める(prev:暫定高(安)値、current:終値)　偽：終値を決める
                    current1 = sheet.cell(row=i,column=3).value
                    n225data = []
                    #list内([0]=日付,[1]=時間,[2]=価格)
                    n225data += [sheet.cell(row=i,column=1).value.date(),sheet.cell(row=i,column=2).value,current1]
                    prev1 = current1
                    current1 = sheet.cell(row=i,column=6).value
                    list1.append(n225data)
                else:
                    current1 = sheet.cell(row=i,column=6).value

                #暫定高(安)値と終値の幅が50円以上の場合
                if abs(current1 - prev1) >= 50:
                    #終値を暫定高(安)値として、日時と値段をlistに追加
                    n225data = []
                    n225data += [sheet.cell(row=i,column=1).value.date(),sheet.cell(row=i,column=2).value,current1]
                    prev1 = current1
                    list1.append(n225data)
            else:
                yori_frag = False

        elif 3 < sheet.cell(row=i,column=1).value.month and sheet.cell(row=i,column=1).value.month <= 6:
            if yori_time == sheet.cell(row=i,column=2).value:
                yori_frag = True
            
            if yori_frag == True and sheet.cell(row=i,column=2).value <= hike_time:
                if yori_time == sheet.cell(row=i,column=2).value and abs(sheet.cell(row=i,column=3).value - prev2) >= 50:
                    current2 = sheet.cell(row=i,column=3).value
                    n225data = []
                    n225data += [sheet.cell(row=i,column=1).value.date(),sheet.cell(row=i,column=2).value,current2] 
                    prev2 = current2
                    current2 = sheet.cell(row=i,column=6).value
                    list2.append(n225data)
                else:
                    current2 = sheet.cell(row=i,column=6).value

                if abs(current2 - prev2) >= 50:
                    n225data = []
                    n225data += [sheet.cell(row=i,column=1).value.date(),sheet.cell(row=i,column=2).value,current2]
                    prev2 = current2
                    list2.append(n225data)
            else:
                yori_frag = False

        elif 6 < sheet.cell(row=i,column=1).value.month and sheet.cell(row=i,column=1).value.month <= 9:
            if yori_time == sheet.cell(row=i,column=2).value:
                yori_frag = True

            if yori_frag == True and sheet.cell(row=i,column=2).value <= hike_time:
                if yori_time == sheet.cell(row=i,column=2).value and abs(sheet.cell(row=i,column=3).value - prev3) >= 50:
                    current3 = sheet.cell(row=i,column=3).value
                    n225data = []
                    n225data += [sheet.cell(row=i,column=1).value.date(),sheet.cell(row=i,column=2).value,current3]
                    prev3 = current3
                    current3 = sheet.cell(row=i,column=6).value
                    list3.append(n225data)
                else:
                    current3 = sheet.cell(row=i,column=6).value

                if abs(current3 - prev3) >= 50:
                    n225data = []
                    n225data += [sheet.cell(row=i,column=1).value.date(),sheet.cell(row=i,column=2).value,current3]
                    prev3 = current3
                    list3.append(n225data)
            else:
                yori_frag = False

        elif 9 < sheet.cell(row=i,column=1).value.month and sheet.cell(row=i,column=1).value.month <= 12:
            if yori_time == sheet.cell(row=1,column=2).value:
                yori_frag = True

            if yori_frag == True and sheet.cell(low=i,column=2).value <= hike_time:
                if yori_time == sheet.cell(row=i,column=2).value and abs(sheet.cell(row=i,column=3).value - prev4) >= 50:
                    current4 = sheet.cell(row=i,column=3).value
                    n225data = []
                    n225data += [sheet.cell(row=i,column=1).value.date(),sheet.cell(row=i,column=2).value,current4]
                    prev4 = current4
                    current4 = sheet.cell(row=i,column=6).value
                    list4.append(n225data)
                else:
                    current4 = sheet.cell(row=i,column=6).value

                if abs(current4 - prev4) >= 50:
                    n225data = []
                    n225data += [sheet.cell(row=i,column=1).value.date(),sheet.cell(row=i,column=2).value,current4]
                    prev4 = current4
                    list4.append(n225data)
            else:
                yori_frag = False
        i += 1

    print("please input the sonnekihyou name ->")
    input_data_kagi = input()
    entry_point1_kagi = []
    entry_point2_kagi = []
    entry_point3_kagi = []
    entry_point4_kagi = []

    kagiSignal(list1,entry_point1_kagi)
    kagiSignal(list2,entry_point2_kagi)
    kagiSignal(list3,entry_point3_kagi)
    kagiSignal(list4,entry_point4_kagi)

    entry_point_kagi = entry_point1_kagi + entry_point2_kagi + entry_point3_kagi + entry_point4_kagi

    make_performance(entry_point_kagi,input_data_kagi+".xlsx")

def kagiSignal(list,entry_point):
    list_count = 0
    retu_count = 1
    #prev_listには暫定高(安)値とトレンド方向を記録,[0]=価格,[1]=トレンド方向
    prev_list = []
    #dequeは先頭、末尾どちらからも取り出せる(キュー、スタック(キューの取り出しはpopleft())
    #yama_tini_queueは直近の山谷形成価格を記録(高値、安値の2つ以上は存在しない)
    yama_tini_queue = deque([])

    yama_nuki = False
    tani_nuki = False

    #for 変数　in --: --の要素を変数として順番にする
    for w in list:
        #list1の要素1-2で上昇か下降か判断、暫定高(安)値をprev_list1に記録
        if list_count == 0 and list[0][2] > list[1][2]:
            prev_list += [list[1][2],-1]
        elif list_count == 0 and list[0][2] < list[1][2]:
            prev_list += [list[1][2],1]

        #記録されたトレンド方向に記録価格が進んでいたら暫定高(安)値更新
        if list_count >= 2:
            if prev_list[1] == 1 and prev_list[0] < w[2]:
                prev_list[0] = w[2]
                prev_list[1] = 1
            elif prev_list[1] == -1 and prev_list[0] > w[2]:
                prev_list[0] = w[2]
                prev_list[1] = -1
            #トレンドと逆方向に記録価格が進んだ場合
            elif prev_list[1] == 1 and prev_list[0] > w[2]:
                #記録されている直近の山谷形成価格が2つある場合、古い記録を削除し、新しい記録に更新(削除されるのは更新される価格と同じ向き)
                #ない場合は追加
                if len(yama_tini_queue) == 2:
                    yama_tini_queue.popleft()
                    yama_tini_queue.append(prev_list)
                else:
                    yama_tini_queue.append(prev_list)

                prev_list = [w[2],-1]

                retu_count += 1

            elif prev_list[1] == -1 and prev_list[0] < w[2]:
                if len(yama_tini_queue) == 2:
                    yama_tini_queue.popleft()
                    yama_tini_queue.append(prev_list)
                else:
                    yama_tini_queue.append(prev_list)
                
                #prev_list1をStructure()で初期化 ⇒ 初期化しないとyama_tini_queue内の値まで変わるから(Excelで関数コピーった時と同じ動作？)
                prev_list = [w[2],1]

                retu_count += 1

            #yama_tini_queueが2つ記録しているとき、前の記録を取り出し現在価格、記録価格、トレンド方向、価格差、リスト数を表示
            if len(yama_tini_queue) == 2:
                temp = yama_tini_queue.popleft()
                print(w[2],temp[0],temp[1],w[2]-temp[0],list_count)

                #山谷超えのデータをtemp3(価格、日付、シグナル)としてnuki_pointに記録
                #超えなかった場合はyama_tini_queueに戻す
                if temp[1] == 1 and w[2] - temp[0] > 0:
                    temp3 = [w[0],w[1],w[2],"L",retu_count]
                    yama_nuki = True
                    entry_point.append(temp3)
                elif temp[1] == -1 and temp[0] - w[2] > 0:
                    temp3 = [w[0],w[1],w[2],"S",retu_count]
                    tani_nuki = True
                    entry_point.append(temp3)
                else:
                    yama_tini_queue.appendleft(temp)
        list_count += 1

def make_performance(entry_point,save_name):
    wb2 =px.Workbook()
    sheet2 = wb2.active
    sheet2.title =  "kagiashi"
    sheet2["A1"] = "エントリー日付"
    sheet2["B1"] = "エントリー"
    sheet2["C1"] = "決済日付"
    sheet2["D1"] = "決済"
    sheet2["E1"] = "シグナル"
    sheet2["F1"] = "損益"
    sheet2["G1"] = "損益合計"

    sum_sonneki = 0
    column_count = 2
    for i in range(1,len(entry_point)):
        if entry_point[i-1][3] == 1:
            sonneki = entry_point[i-1][2] - entry_point[i][2]
        else:
            sonneki = entry_point[i][2] - entry_point[i-1][2]

        sum_sonneki += sonneki*100

        sheet2.cell(row=column_count, column=1, value=entry_point[i-1][0])
        sheet2.cell(row=column_count, column=2, value=entry_point[i-1][2])
        sheet2.cell(row=column_count, column=3, value=entry_point[i][0])
        sheet2.cell(row=column_count, column=4, value=entry_point[i][2])
        sheet2.cell(row=column_count, column=5, value=entry_point[i-1][3])
        sheet2.cell(row=column_count, column=6, value=sonneki)
        sheet2.cell(row=column_count, column=7, value=sum_sonneki)

        column_count += 1

    wb2.save(save_name)

get225Data()
