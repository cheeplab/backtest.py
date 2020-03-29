#実運用のルールのバックテスト(excel.csvデータを入力、出力)
#ルール　①Hyperによる取引(日中と夜の立会ごとに決済）　②値幅30円で更新　③前の山(谷)を超えて+30で新規、前の谷(山)を超えて-20で決済　④含み損100で損切り　⑤手数料は売買1回で100円と仮定 ⑥1年分でテスト(sq日ごとに計算はしない)
import os
import pandas as pd
import openpyxl as px
from openpyxl.chart import LineChart,Reference,series
import datetime
from collections import deque

#df_2 = [日付、時間、始値、高値、安値、終値](日付、時間はdatetime型(時間をtime型にしたかったけどよく分からなかった))
#夜相場の閉め時のデータは削除、次の関数でindexが倍数になるごとにリストを作るためにキリ良くしたかった
def get1minsheet():
    time_1 = datetime.datetime(1900,1,1,5,30)
    time_2 = datetime.datetime(1900,1,1,15,15)

    df_1 = pd.read_csv(os.getcwd()+"\\N225minif_2019.csv",skiprows=1,names=("Date","Time","Open","High","Low","Close"),encoding='cp932',usecols=["Date","Time","Open","High","Low","Close"])

    df_1["Date"] = pd.to_datetime(df_1["Date"],format='%Y/%m/%d')
    df_1["Time"] = pd.to_datetime(df_1["Time"],format='%H:%M')

    df_2 = df_1[(df_1["Time"] != time_1)&(df_1["Time"] != time_2)]
    
    return(df_2.reset_index(drop=True))

#ohlc_5 = [日付、時間、始値、[[高値、安値](1分足)*tf]、終値]⇒([0]:日付,[1]:時間,[2]:始値,[3]([0],[1]):高値、安値,[4]:終値)
def make_ohlc(df,tf):
    ohlc = []
    hl_1 = []

    #append:１つの要素を追加可能、複数加える場合は2次元内包になる　extend:1度に複数要素を追加可能
    for num in range(len(df.index)-1):
        if num % tf == 0:
            ohlc_5 = []
            hl_1 = []
            ohlc_5.extend([df.iat[num,0],df.iat[num,1],df.iat[num,2]])

        hl_1.append([df.iat[num,3],df.iat[num,4]])
        
        if num % tf == tf-1:
            ohlc_5.append(hl_1)
            ohlc_5.append(df.iat[num,5])
            ohlc.append(ohlc_5)

    return(ohlc)

#日中立会内での暫定高(安)値とその時の日時を記録する関数
def get225Data():
    
    print("作成したい足を入力してください　ex)1H ⇒ 60")
    input_min = input()
    input_min = int(input_min)
    sheet = make_ohlc(get1minsheet(),input_min)

    #[]:listを初期化する
    #yama,tani,entryは初期値として絶対超えない値を置いた
    data = []
    prev = 0
    yama = 100000
    tani = 0
    entry_L = 100000
    entry_S = 0
    frag = 0
    trend = 0
    yori_time_am = datetime.datetime(1900,1,1,8,45) 
    hike_time_am = datetime.datetime(1900,1,1,15,10)
    yori_time_pm = datetime.datetime(1900,1,1,16,30)
    hike_time_pm = datetime.datetime(1900,1,1,5,25)

    #sheetに要素がなくなるまで繰り返す関数
    for i in sheet:
        #相場開始時刻ならば相場ごとのリストを初期化して始値と終値を比較、始値をprevに設定
        if i[1] == yori_time_am or i[1] == yori_time_pm:
            per_data = []
            prev = i[2]

            #prevを各終値が超えていた場合カギ足作成⇒trendが-1ならば次は谷を、1なら次は山を作成する
            #山谷作成と同時に+30にエントリーポイントを作る、各分足で確認
            if abs(i[4]-i[2]) >= 30:
                if i[4] < prev:
                    yama = prev
                    entry_L = yama+30
                    trend = -1
                else:
                    tani = prev
                    entry_S = tani-30
                    ternd = 1

                prev = i[4]

        elif i[1] <= hike_time_am or yori_time_pm < i[1]:
            #分足での高(安)値がエントリーor決済に入ってるか調べる。fragが立っていれば決済、なければエントリーポイントを探す
            #決済時の判定優先は(損切り)>(利確)
            for j in range(input_min):
                #エントリー-100 or 谷-20を安値が下回ったら決済、相場ごとのリストに[~、決済時間、決済価格、シグナル]を記載、全体リストに追加
                if frag == 1:
                    if entry_L-100 > i[3][j][1]: 
                        per_data.extend([i[1],entry_L-100,"L"])
                        data.append(per_data)
                        per_data = []
                        frag = 0
                    elif tani-20 > i[3][j][1]:
                        per_data.extend([i[1],tani-20,"L"])
                        data.append(per_data)
                        per_data = []
                        frag = 0
                elif frag == -1:
                    if entry_S+100 < i[3][j][0]:
                        per_data.extend([i[1],entry_S+100,"S"])
                        data.append(per_data)
                        per_data = []
                        frag = 0
                    elif yama+20 < i[3][j][0]:
                        per_data.extend([i[1],yama+20,"S"])
                        data.append(per_data)
                        per_data = []
                        frag = 0
                #エントリーを高値(安値)が上回ったらエントリー、フラッグを立てて分足で決済を探す⇒相場ごとのリストに[日付、エントリー時間、エントリーポイント]を記載
                else:
                    if i[3][j][0] > entry_L:
                        per_data.extend([i[0],i[1],entry_L])
                        frag = 1
                    elif i[3][j][1] < entry_S:
                        per_data.extend([i[0],i[1],entry_S])
                        frag = -1

            if i[1] == hike_time_am or i[1] == hike_time_pm:
                if frag == 1:
                    per_data.extend([i[1],i[4],"L"])
                    data.append(per_data)
                    per_data = []
                    frag = 0
                elif frag == -1:
                    per_data.extend([i[1],i[4],"S"])
                    data.append(per_data)
                    per_data = []
                    frag = 0
                
            #trend方向と値幅の正負が一緒の時はprevを更新するだけ
            elif abs(i[4]-prev) >= 30:
                if i[4] > prev:
                    if trend == 1:
                        prev = i[4]
                    else:
                        tani = prev
                        entry_S = tani-30
                        trend = 1
                else:
                    if trend == -1:
                        prev = i[4]
                    else:
                        yama = prev
                        entry_L = yama+30
                        trend =-1

    print("損益表名を入力してください ->")
    input_data_kagi = input()

    make_performance(data,input_data_kagi+".xlsx")


def make_performance(entry_point,save_name):
    wb2 =px.Workbook()
    sheet2 = wb2.active
    sheet2.title =  "kagiashi"
    sheet2["A1"] = "エントリー日付"
    sheet2["B1"] = "エントリー時間"
    sheet2["C1"] = "決済時間"
    sheet2["D1"] = "エントリー"
    sheet2["E1"] = "決済"
    sheet2["F1"] = "シグナル"
    sheet2["G1"] = "損益"
    sheet2["H1"] = "損益合計"

    sum_sonneki = 0
    column_count = 2
    #sonneki内の-1は手数料代100円を想定
    for k in entry_point:
        if k[5] == 'L':
            sonneki = (k[4]-k[2]-1)*100
        else:
            sonneki = (k[2]-k[4]-1)*100

        sum_sonneki += sonneki

        sheet2.cell(row=column_count, column=1, value=k[0])
        sheet2.cell(row=column_count, column=2, value=k[1])
        sheet2.cell(row=column_count, column=3, value=k[3])
        sheet2.cell(row=column_count, column=4, value=k[2])
        sheet2.cell(row=column_count, column=5, value=k[4])
        sheet2.cell(row=column_count, column=6, value=k[5])
        sheet2.cell(row=column_count, column=7, value=sonneki)
        sheet2.cell(row=column_count, column=8, value=sum_sonneki)

        column_count += 1

    #折れ線グラフの作成
    values = Reference(sheet2,min_col=8,min_row=2,max_col=8,max_row=len(entry_point)+1)
    chart = LineChart()
    chart.add_data(values)
    sheet2.add_chart(chart,"I2")

    wb2.save(save_name)